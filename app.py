# app.py

from flask import Flask, render_template, request, redirect, url_for, flash
import re
import dns.resolver
import smtplib
import openpyxl
from concurrent.futures import ThreadPoolExecutor
import os

app = Flask(__name__)
app.secret_key = 'your_secret_key'  # Replace with a strong secret key

# Function to validate email syntax
def validate_email_syntax(email):
    pattern = r"^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$"
    return bool(re.match(pattern, email))

# Function to validate domain
def validate_domain(email):
    try:
        domain = email.split('@')[1]
        dns.resolver.resolve(domain, 'A')
        return True
    except (dns.resolver.NXDOMAIN, dns.resolver.NoAnswer):
        return False

# Function to check MX record
def check_mx_record(email):
    try:
        domain = email.split('@')[1]
        mx_records = dns.resolver.resolve(domain, 'MX')
        return bool(mx_records)
    except (dns.resolver.NXDOMAIN, dns.resolver.NoAnswer):
        return False

# Function to verify email via SMTP
def smtp_verify(email):
    domain = email.split('@')[1]
    try:
        mx_records = dns.resolver.resolve(domain, 'MX')
        mx_record = str(mx_records[0].exchange)
        server = smtplib.SMTP(timeout=10)
        server.connect(mx_record)
        server.helo(server.local_hostname)
        server.mail('test@example.com')
        code, message = server.rcpt(email)
        server.quit()
        return code == 250
    except Exception as e:
        print(f"Failed to verify {email}: {e}")
    return False

# Function to determine if an email is deliverable
def is_email_deliverable(email):
    if not validate_email_syntax(email):
        return False
    if not validate_domain(email):
        return False
    if not check_mx_record(email):
        return False
    if not smtp_verify(email):
        return False
    return True

# Route for the home page
@app.route('/')
def index():
    return render_template('index.html')

# Route for single email verification
@app.route('/single_verify', methods=['GET', 'POST'])
def single_verify():
    if request.method == 'POST':
        email = request.form['email']
        if is_email_deliverable(email):
            flash(f"Email {email} is deliverable.", "success")
        else:
            flash(f"Email {email} is not deliverable.", "danger")
        return redirect(url_for('single_verify'))
    return render_template('single_verify.html')

# Function to process bulk emails
def process_email(email, total_emails, counter):
    result = email if is_email_deliverable(email) else None
    counter[0] += 1
    return result

# Route for bulk email verification
@app.route('/bulk_verify', methods=['GET', 'POST'])
def bulk_verify():
    if request.method == 'POST':
        file = request.files['file']
        if file and file.filename.endswith('.xlsx'):
            file_path = os.path.join('uploads', file.filename)
            file.save(file_path)

            email_list = read_emails_from_excel(file_path, "Sheet1")
            total_emails = len(email_list)
            counter = [0]

            with ThreadPoolExecutor(max_workers=10) as executor:
                results = list(executor.map(lambda email: process_email(email, total_emails, counter), email_list))

            deliverable_emails = [email for email in results if email is not None]
            write_emails_to_excel('deliverable_emails.xlsx', deliverable_emails)

            flash("Bulk verification completed. Check deliverable_emails.xlsx for results.", "success")
        else:
            flash("Please upload a valid .xlsx file.", "danger")
        return redirect(url_for('bulk_verify'))
    return render_template('bulk_verify.html')

# Function to read emails from Excel
def read_emails_from_excel(file_path, sheet_name):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[sheet_name]
    emails = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        email = row[0]
        if email:
            emails.append(email)
    return emails

# Function to write deliverable emails to Excel
def write_emails_to_excel(file_path, emails, sheet_name="Deliverable Emails"):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = sheet_name
    sheet.append(["Email", "Status"])
    for email in emails:
        sheet.append([email, "Deliverable"])
    wb.save(file_path)

if __name__ == '__main__':
    app.run(debug=True)
